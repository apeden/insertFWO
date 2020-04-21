from openpyxl import load_workbook as load
import csv

#columns in csv file, counting from zero
date_col = 2
name_col = 6   
amount_col = 7  #amounts in csv file
code_col = 3

max_weeks_in_year = 55
ignore_terms = ("murray","glen", "wood", "cancel", "cash", "deleted", "clearance", "0")
DDdict = {}

def ignore(list_of_names, ignore_terms):
    for term in ignore_terms :
        for name_part in list_of_names:
            if term in name_part:
                return True
    return False

workbook = "../2019/Gift Aid Analysis 2019_blank.xlsx"
csv_file = "../2019/AlexForExport02_01_20_ABC"

try:
    wb = load(workbook, data_only = True)
except PermissionError:
    print("You currently have the excel workbook open")
    raise


"""search through excel names for match in name from sage"""
def xl_nm_indx_fndr(sage_name_parts, sheet, num_excel_names = 82):
    name_index = 0
    while name_index <= num_excel_names:                
        excel_name = str(sheet.cell(column = 3, row = (name_index + 5)).value)               
            #attempts to match name in excel file...
            #with name in csv (all lower case) 
        for name_part in sage_name_parts:
            if excel_name.lower() == name_part:
                return name_index
        name_index += 1
    
def write_amount_for_month(row, sheet, excel_name_index):
    month = int(row[date_col][3:5])
    amountPaid = float(row[amount_col])
    try:
        DDdict[(row[date_col][3:5],excel_name_index)] += amountPaid
    except:
        DDdict[(row[date_col][3:5],excel_name_index)] = amountPaid        
    print(row[name_col],row[date_col],amountPaid,": ",excel_name_index)
                #adds amount to existing amount in appropriate cell
    sheet.cell(column =(5 + month),
                row = (excel_name_index + 5),
                value = DDdict[(row[date_col][3:5],excel_name_index)])
    return DDdict[(row[date_col][3:5],excel_name_index)]

def insertDD(sheet, csv_file, ignore_terms, output_file):
    amount_total = 0.0
    name_index = None
    with open(csv_file+".csv") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
                #goes down csv file row by row   
        for row in csv_reader:
            sage_name_parts = row[name_col].lower().split()
            if ignore(sage_name_parts, ignore_terms):
                continue
                #passes over account entries already covered in cash...
                #and where is says cancel, cash, deleted etc
            excel_name_index = xl_nm_indx_fndr(sage_name_parts, sheet)
            try:
                amount_total += write_amount_for_month(row,
                                                       sheet,
                                                       excel_name_index)
            except:
                print(row[name_col], " not found in list of names in excel file")
    wb.save(output_file)
    return amount_total
    
def insertCash(sheet,csv_file):   
                #set Sage entry in excel file to None
    for i in range(max_weeks_in_year):
        sheet.cell(column =(6 + i), row = 92, value = None)        
    with open(csv_file+".csv") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
                #goes down csv file row by row
        for row in csv_reader:
            for i in range(max_weeks_in_year):
                #looks at all the codes in the excel sheet
                code = str(sheet.cell(column = (6 + i), row = 1).value)
                #if code from excel is code in row of csv file....
                full_name = row[name_col].split()
                if "Cancel" not in full_name and\
                   "Deleted" not in full_name and\
                   "Clearance of year end prepayments" not in full_name:
                    if code == row[code_col] or \
                       code == row[code_col].replace(" ",""):                    
                        cash = float(row[amount_col]) 
                        currCellValue = 0.0
                #existing amount in the distination cell
                        valStr = (str
                                  (sheet.cell
                                   (column = (6 + i), row = 92).value))
                        if valStr != "None":
                            currCellValue = float(valStr)
                #adds amount to existing amount in appropriate cell
                        sheet.cell(column =(6 + i),
                                  row = 92,
                                  value = (cash + currCellValue))
                #copy of excel file saved with new name
    wb.save("Gift Aid Analysis 2019 CASH copy.xlsx")

def findTrans(csv_file, *codes):
   with open(csv_file+".csv") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
                #goes down csv file row by row
        for row in csv_reader:
            for code in codes:
                if code == row[code_col] or code == row[code_col].replace(" ",""):
                    print(row[code_col],", ",
                          row[date_col],", ",
                          row[name_col],", ",
                          row[amount_col])
                    
insertCash(wb["envelopes"], "../2019/FWO 2019-Final")
##print(insertDD(wb["ABC"],
##               csv_file,
##               ("day-",),
##               "Gift Aid Analysis 2019 ABC copy.xlsx"))
##print(insertDD(wb["DONATIONS"],
##               "../2019/AlexForExport02_01_20_DON",
##               (),
##               "Gift Aid Analysis 2019 DON copy.xlsx"))
##print(insertDD(wb["DD"],
##               "../2019/FWO 2019-Final",
##               ignore_terms,
##               "Gift Aid Analysis 2019 DD copy new.xlsx"))
##findTrans(csv_file, "BR17")
